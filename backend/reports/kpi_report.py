"""
Inventory KPI Reporting Service
================================
Generates per-worker daily KPI reports from MongoDB transaction data.

Usage:
    python kpi_report.py                     # Today's report (console)
    python kpi_report.py --date 2026-04-10   # Specific date
    python kpi_report.py --range 7           # Last 7 days
    python kpi_report.py --csv               # Export to CSV
    python kpi_report.py --xlsx              # Export to Excel
    python kpi_report.py --worker AAMIR      # Filter by worker

Requires:
    pip install pymongo python-dotenv openpyxl
"""

import os
import sys
import argparse
import csv
from datetime import datetime, timedelta
from pathlib import Path

try:
    from pymongo import MongoClient
except ImportError:
    print("ERROR: pymongo not installed. Run: pip install pymongo")
    sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

# ── Load environment ─────────────────────────────────────────────────────────
env_path = Path(__file__).resolve().parent / ".env"
if load_dotenv and env_path.exists():
    load_dotenv(env_path)

MONGO_URI = os.environ.get(
    "MONGODB_URI",
    os.environ.get("MONGO_URI", ""),
)

if not MONGO_URI:
    print("ERROR: MONGODB_URI not set. Add it to backend/.env or environment.")
    sys.exit(1)


# ── Database connection ──────────────────────────────────────────────────────
def get_db():
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=10000)
    db_name = MONGO_URI.rsplit("/", 1)[-1].split("?")[0] or "inventory"
    return client[db_name]


# ── Report: Per-Worker Daily KPI ─────────────────────────────────────────────
def worker_daily_kpi(db, target_date, worker_filter=None):
    """
    For each worker on the given date, compute:
      - total_transactions: how many rows they synced
      - pending: rows still in pending status
      - processed: rows marked processed
      - archived: rows marked archived
      - unique_items: distinct barcodes scanned
      - first_scan: earliest Timestamp
      - last_scan: latest Timestamp
      - active_hours: difference between first and last scan
      - bins_touched: unique (Frombin, Tobin) pairs
    """
    start = datetime.combine(target_date, datetime.min.time())
    end = start + timedelta(days=1)

    match_stage = {
        "Timestamp": {"$gte": start, "$lt": end},
    }
    if worker_filter:
        match_stage["Worker_Name"] = worker_filter

    pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$Worker_Name",
                "total_transactions": {"$sum": 1},
                "total_qty": {"$sum": {"$ifNull": ["$Qty", 0]}},
                "pending": {
                    "$sum": {
                        "$cond": [
                            {
                                "$or": [
                                    {"$eq": ["$syncStatus", "pending"]},
                                    {"$eq": ["$syncStatus", None]},
                                    {
                                        "$not": {
                                            "$ifNull": ["$syncStatus", False]
                                        }
                                    },
                                ]
                            },
                            1,
                            0,
                        ]
                    }
                },
                "processed": {
                    "$sum": {
                        "$cond": [{"$eq": ["$syncStatus", "processed"]}, 1, 0]
                    }
                },
                "archived": {
                    "$sum": {
                        "$cond": [{"$eq": ["$syncStatus", "archived"]}, 1, 0]
                    }
                },
                "unique_items": {"$addToSet": "$Item_Barcode"},
                "first_scan": {"$min": "$Timestamp"},
                "last_scan": {"$max": "$Timestamp"},
                "from_bins": {"$addToSet": "$Frombin"},
                "to_bins": {"$addToSet": "$Tobin"},
            }
        },
        {"$sort": {"total_transactions": -1}},
    ]

    results = list(db.transactions.aggregate(pipeline))

    rows = []
    for r in results:
        worker = r["_id"] or "unknown"
        first = r.get("first_scan")
        last = r.get("last_scan")
        active_mins = 0
        if first and last:
            active_mins = round((last - first).total_seconds() / 60)

        all_bins = set(r.get("from_bins", [])) | set(r.get("to_bins", []))
        all_bins.discard("")
        all_bins.discard(None)

        rows.append(
            {
                "date": target_date.isoformat(),
                "worker": worker,
                "total_transactions": r["total_transactions"],
                "total_qty": r.get("total_qty", 0),
                "pending": r["pending"],
                "processed": r["processed"],
                "archived": r["archived"],
                "unique_items": len(r.get("unique_items", [])),
                "bins_touched": len(all_bins),
                "first_scan": first.strftime("%H:%M:%S") if first else "",
                "last_scan": last.strftime("%H:%M:%S") if last else "",
                "active_minutes": active_mins,
            }
        )

    return rows


# ── Report: Summary Totals ───────────────────────────────────────────────────
def summary_totals(db):
    """Quick snapshot of overall queue health."""
    pending = db.transactions.count_documents(
        {
            "$or": [
                {"syncStatus": {"$exists": False}},
                {"syncStatus": None},
                {"syncStatus": "pending"},
            ]
        }
    )
    processed = db.transactions.count_documents({"syncStatus": "processed"})
    archived = db.transactions.count_documents({"syncStatus": "archived"})
    total_items = db.items.count_documents({})

    return {
        "total_items": total_items,
        "pending": pending,
        "processed": processed,
        "archived": archived,
        "grand_total": pending + processed + archived,
    }


# ── Output: Console Table ────────────────────────────────────────────────────
def print_table(rows, title=""):
    if title:
        print(f"\n{'='*70}")
        print(f"  {title}")
        print(f"{'='*70}")

    if not rows:
        print("  (no data)")
        return

    headers = list(rows[0].keys())
    col_widths = {h: max(len(h), max(len(str(r.get(h, ""))) for r in rows)) for h in headers}

    header_line = " | ".join(h.ljust(col_widths[h]) for h in headers)
    print(f"  {header_line}")
    print(f"  {'-' * len(header_line)}")

    for row in rows:
        line = " | ".join(str(row.get(h, "")).ljust(col_widths[h]) for h in headers)
        print(f"  {line}")

    print()


# ── Output: CSV ──────────────────────────────────────────────────────────────
def export_csv(rows, filename):
    if not rows:
        print(f"  No data to export to {filename}")
        return
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Exported {len(rows)} rows to {filename}")


# ── Output: Excel ────────────────────────────────────────────────────────────
def export_xlsx(rows, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    if not rows:
        print(f"  No data to export to {filename}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Report"

    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(r.get(header, ""))) for r in rows), default=0))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    wb.save(filename)
    print(f"  Exported {len(rows)} rows to {filename}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Inventory KPI Report Generator")
    parser.add_argument("--date", help="Target date (YYYY-MM-DD). Default: today")
    parser.add_argument("--range", type=int, help="Generate report for last N days")
    parser.add_argument("--worker", help="Filter by worker name")
    parser.add_argument("--csv", action="store_true", help="Export to CSV file")
    parser.add_argument("--xlsx", action="store_true", help="Export to Excel file")
    parser.add_argument("--summary", action="store_true", help="Show queue summary only")
    args = parser.parse_args()

    db = get_db()

    # Summary mode
    if args.summary:
        totals = summary_totals(db)
        print("\n  Queue Summary")
        print(f"  {'─'*40}")
        for k, v in totals.items():
            print(f"  {k.replace('_', ' ').title():.<30} {v}")
        print()
        return

    # Determine date range
    if args.date:
        target = datetime.strptime(args.date, "%Y-%m-%d").date()
        dates = [target]
    elif args.range:
        today = datetime.now().date()
        dates = [today - timedelta(days=i) for i in range(args.range)]
        dates.reverse()
    else:
        dates = [datetime.now().date()]

    # Collect all rows
    all_rows = []
    for d in dates:
        rows = worker_daily_kpi(db, d, args.worker)
        all_rows.extend(rows)

    # Output
    title = f"Worker KPI Report"
    if args.worker:
        title += f" — {args.worker}"
    if len(dates) == 1:
        title += f" — {dates[0].isoformat()}"
    else:
        title += f" — {dates[0].isoformat()} to {dates[-1].isoformat()}"

    if args.csv:
        date_str = dates[0].isoformat() if len(dates) == 1 else f"{dates[0].isoformat()}_to_{dates[-1].isoformat()}"
        filename = f"kpi_{date_str}.csv"
        export_csv(all_rows, filename)
    elif args.xlsx:
        date_str = dates[0].isoformat() if len(dates) == 1 else f"{dates[0].isoformat()}_to_{dates[-1].isoformat()}"
        filename = f"kpi_{date_str}.xlsx"
        export_xlsx(all_rows, filename)
    else:
        print_table(all_rows, title)

    # Always show summary
    totals = summary_totals(db)
    print(f"  Queue: {totals['pending']} pending | {totals['processed']} processed | {totals['archived']} archived | {totals['total_items']} items")


if __name__ == "__main__":
    main()
